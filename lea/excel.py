import time
from threading import Thread

import openpyxl as openpyxl


def opex():
    wb = openpyxl.load_workbook("../file/问题.xlsx")
    print(wb.sheetnames)
    sheet = wb['Sheet1']

    for row in sheet.rows:
        row_str = ""
        for cell in row:
            cell_str = cell.value
            print(cell_str, end=",")
        print()


def wtex():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(1, 1, "哈哈哈哈")
    sheet.cell(1, 2, "哈哈哈哈2")
    sheet.cell(1, 3, "哈哈哈哈3")
    sheet.cell(2, 1, "张三")
    sheet.cell(2, 2, "李四")
    sheet.cell(2, 3, "玩")
    wb.save(filename="test.xlsx")


def saveFile(wb, name):
    wb.save(filename=name)


def xx():
    with open("../file/data.txt", 'r', encoding='utf-8') as f:
        dict = {}
        while True:
            line = f.readline()
            if not line:
                break
            ss = line.split("	")
            key = ss[0]
            val = ss[1].strip()
            if dict.get(key) is None:
                dict[key] = []
            dict[key].append(val)
        for key in dict.keys():
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(1, 1, "姓名")
            sheet.cell(1, 2, "电话")
            index = 2
            for value in dict[key]:
                sheet.cell(index, 2, value)
                index = index + 1
            name = "C:/Users/ll18/Desktop/xxxx/xx2/" + key + ".xlsx"
            wb.save(filename=name)
            # thread = Thread(target=saveFile(wb,name))
            # thread.start()
            # thread.join()


if __name__ == '__main__':
    s = time.time()
    xx()
    e = time.time()
    print("耗时 %s" % str((e-s)/1000))
